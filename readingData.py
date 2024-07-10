import config.database as database
from openpyxl import Workbook
import os

class_names = ['husky', 'beagle', 'rottweiler', 'german-shepherd', 'dalmatian', 'poodle', 'bulldog', 'labrador-retriever']

def get_image_data(folder_path,breed):
  image_data = []
  for filename in os.listdir(folder_path):
    if filename.lower().endswith((".jpg", ".jpeg", ".png")):
      image_path = os.path.join(folder_path, filename)
      image_data.append({"name":filename,"path":image_path, "breed":breed})
  return image_data

def write_to_excel(image_data, excel_filename):
  workbook = Workbook()
  worksheet = workbook.active
  worksheet.cell(row=1, column=1).value = "name"
  worksheet.cell(row=1, column=2).value = "path"
  worksheet.cell(row=1, column=3).value = "breed"

  row_num = 2
  for image_name, image_path, breed in image_data:
    worksheet.cell(row=row_num, column=1).value = image_name
    worksheet.cell(row=row_num, column=2).value = image_path
    worksheet.cell(row=row_num, column=3).value = breed
    row_num += 1

  workbook.save(excel_filename)

if __name__ == "__main__":
  for breed in class_names:
    petCollection=database.db.get_collection("pets")
    folder_path = "./dog-breeds/"+breed
    excel_filename = "dog_data.xlsx"
    animal_data = get_image_data(folder_path, breed)
    petCollection.insert_many(animal_data)
